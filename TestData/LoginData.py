
import openpyxl

class LoginData:
    test_Login_data = [{"username": "standard_user", "password": "secret_sauce"},
                    {"username": "visual_user", "password": "secret_sauce"}]

    @staticmethod
    def getTestData():  # self parameter not required for static method
        data_list = []
        book = openpyxl.load_workbook(
            "/home/nayana/PycharmProjects/DemoSauceLabProject/TestData/LoginDataSource.xlsx")
        sheet = book.active
        for i in range(2, sheet.max_row + 1):
            data_dict = {}
            for j in range(2, sheet.max_column + 1):
                data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            data_list.append(data_dict)

        return data_list
